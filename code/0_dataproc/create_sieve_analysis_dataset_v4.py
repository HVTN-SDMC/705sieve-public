"""

Program: create_sieve_analysis_dataset_v4.py

Purpose: Create the sieve analysis dataset for HVTN 705 final analysis.
         This is for version 3 of the final sequence files.

Inputs: Sequence file
        Reference sequence file
        Output file

Outputs: HVTN_705_sieve_data_final_v3.csv

Execution: python create_sieve_analysis_dataset.py

Requirements: Python 3 virtual environment with Biopython and pandas.

"""

import argparse
from csv import reader
import re
from Bio import SeqIO
from openpyxl import load_workbook
import pandas as pd

adcc_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/site_lists/v705_adcc_epitope.dat"
ant1428_pos_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/site_lists/v705_1428_V1V2_sites.dat"
antigen1428_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/seq/v2_antigens_lanl_hivalign.fasta"
antigen1428_map_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/seq/map/v705_antigens_final.map"
charge_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/from_egiorgi/v705_env_var_reg_char_for_sieve.xlsm"
cclade_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/site_lists/v705_clade_c_ab.dat"
cysteine_conserved_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/site_lists/v705_sites_cysteine_conserved.dat"
founder_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/from_egiorgi/V705_REN_LineageSummaryResults_Final2023Sieve.xlsm"
hamming_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/zspace/wold_z5_space_diff_v3.csv"
hvtn505_site_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/site_lists/adecamp/HVTN505_signature_sets.csv"
ptid_lookup_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/lookup/v705_ptid_seqid_lookup.csv"
subtype_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/seq/subtype/v705_subtype_env_v2.csv"
survival_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/survival/v705_survival_pp_m_7_24_tau_v3.csv"
unalignable_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/site_lists/v705_unalignable_sites_final.dat"
v1v2_file = "/bioinfo/sponsors/nih/hvtn/hvtn705/sieve/analysis/cmagaret/07_mksieve_final/dat/site_lists/V1V2_correlates_sites_gp70.csv"

ANTIGEN_1428_DESIG = "gp70-001428.2.42"
CLADE_SEQ_DESIG = "CON_C_Env_LANL_2021"
INSERT1_SEQ_DESIG = "Ad26.Mos1.Env"
INSERT2_SEQ_DESIG = "Ad26.Mos2S.Env"
INSERT3_SEQ_DESIG = "C97ZA_Clade_C"
REF_SEQ_DESIG = "HXB2"

ENV_START_POS = 1
ENV_END_POS = 856
GP120_START_POS = 31
GP120_END_POS = 511
GP41_START_POS = 512
GP41_INSERT1_END_POS = 681
GP41_INSERT2_END_POS = 712
GP41_INSERT3_END_POS = 683
INSERT1_START_POS = 1
INSERT1_END_POS = 681
INSERT2_START_POS = 1
INSERT2_END_POS = 712
INSERT3_START_POS = 1
INSERT3_END_POS = 683
V2_START_POS = 157
V2_END_POS = 184
V5_START_POS = 460
V5_END_POS = 470
V1V2_START_POS = 131
V1V2_END_POS = 196

a1428_abs_map_dict = {}
a1428_map_dict = {}
adcc_list = []
antigen1428_list = []
c_ab_list = []
conserved_cysteine_abs_pos_list = []
cysteine_count_list = []
hamming_list = []
hxb2_abs_map_dict = {}
hxb2_abs_sort_map_dict = {}
hxb2_map_dict = {}
length_list = []
non_conserved_cysteine_count_list = []
reference_independent_list = []
region_sequon_list = []
seq_dict = {}
seq_names_list = []
seq_match_mos1_list = []
seq_match_mos2_list = []
seq_match_c97za_list = []
seq_match_1428_list = []
sequon_abs_pos_list = []
sequon_list = []
v2_ab_list = []

parser = argparse.ArgumentParser()

parser.add_argument("seq_file", type=argparse.FileType("r"),
                    help="Sequence file")
parser.add_argument("ref_file", type=argparse.FileType("r"),
                    help="Reference sequence file")
parser.add_argument("hxb2_file", type=str,
                    help="HXB2 sequence file")
parser.add_argument("out_file", type=argparse.FileType("w"),
                    help="Output file")

args = parser.parse_args()

# Read in the survival data
survival_df = pd.read_csv(survival_file)

# Read the PTID/lab seq ID lookup table and merge it with the survival data.
labseqID_map_df = pd.read_csv(ptid_lookup_file)
labseqID_map_df = labseqID_map_df.rename(columns={"ptid": "subjid", "seqid": "seq.id"})

output_df = pd.merge(survival_df, labseqID_map_df, how="left", on="subjid")

# Read in the reference sequences and convert the insert sequences into lists
# to make it easier to compare later.
for seq_rec in SeqIO.parse(args.ref_file, "fasta"):
    if seq_rec.id == REF_SEQ_DESIG:
        continue
    elif seq_rec.id == CLADE_SEQ_DESIG:
        continue
    elif seq_rec.id == INSERT1_SEQ_DESIG:
        insert1_seq = list(str(seq_rec.seq))
    elif seq_rec.id == INSERT2_SEQ_DESIG:
        insert2_seq = list(str(seq_rec.seq))
    elif seq_rec.id == INSERT3_SEQ_DESIG:
        insert3_seq = list(str(seq_rec.seq))

# Read in the subject sequences and convert the insert sequences into lists to
# make it easier to compare later.
for seq_rec in SeqIO.parse(args.seq_file, "fasta"):
    # The sequence ID in the file looks like:
    # V705_num_visit_Env...
    seq_id_parts = seq_rec.id.split("_")
    seq_id = f"{seq_id_parts[0]}_{seq_id_parts[1]}"
    visit = seq_id_parts[2]
    if "-" not in visit:
        visitno = int(visit)
    else:
        visitno = int(visit.split("-")[0])

    if seq_id not in seq_dict:
        # Create a list of dictionaries. The dictionaries contain the sequence
        # ID (derived from the sequence name) and the sequence name.
        seq_name_dict = {}
        seq_name_dict["seq.id"] = seq_id
        seq_name_dict["seq.name"] = seq_rec.id
        seq_name_dict["visit"] = visitno
        seq_names_list.append(seq_name_dict)

        # Create a dictionary of sequences.
        seq_dict[seq_id] = list(str(seq_rec.seq))

    # Some subjects have more than one sequence, so take the earliest one.
    else:
        for seq_list_rec in seq_names_list:
            if (seq_list_rec["seq.id"] == seq_id) and (seq_list_rec["visit"] > visitno):
                seq_list_rec["seq.name"] = seq_rec.id
                seq_list_rec["visit"] = visitno
                seq_dict[seq_id] = list(str(seq_rec.seq))
                break

# Add the sequence name to the output. If the sequence name is missing,
# meaning that there is no sequence available, delete the sequence ID.
seq_names_df = pd.DataFrame.from_records(seq_names_list)
output_df = pd.merge(output_df, seq_names_df, how="left", on="seq.id")
output_df.loc[output_df["seq.name"].isnull(), "seq.id"] = None

# Read in the 1428 V1V2 antigen sequence.
for seq_rec in SeqIO.parse(antigen1428_file, "fasta"):
    if seq_rec.id == ANTIGEN_1428_DESIG:
        insert4_seq = list(str(seq_rec.seq))

# Read in the 1428 antigen coordinate map.
with open(antigen1428_map_file, "r") as a1428mf:
    file_lines = reader(a1428mf, delimiter="|")
    skip_header = next(file_lines)

    if skip_header is not None:
        for coordinates in file_lines:
            # The coordinate map dictionary exists in two forms:
            # 1) with the relative position as the key, in order to mark the
            #    unalignable positions in the insert sequences;
            # 2) with the absolute position as the key in order to create the
            #    correct variable name with the  sequence positions are matched.
            a1428_map_dict[coordinates[1]] = int(coordinates[0]) - 1
            a1428_abs_map_dict[int(coordinates[0]) - 1] = coordinates[1]

# Read in the information on the electrochemical charge of the V2 region. Due
# to the variable order, this data gets merged into the output near the end of
# the program.
charge_wb = load_workbook(filename=charge_file)
charge_ws = charge_wb.active
charge_dict_list = []
for i in range(4, charge_ws.max_row + 1):
    charge_dict = {}
    # The sequence ID in the file looks like:
    # V705_num_visit_Env...
    seq_id_parts = charge_ws[f"A{i}"].value.split("_")
    seq_id = f"{seq_id_parts[0]}_{seq_id_parts[1]}"
    visit = seq_id_parts[2]
    if "-" not in visit:
        visitno = int(visit)
    else:
        visitno = int(visit.split("-")[0])

    if ((seq_names_df["seq.id"] == seq_id) & (seq_names_df["visit"] == visitno)).any():
        charge_dict["seq.id"] = seq_id
        charge_dict["charge.v2"] = str(charge_ws[f"G{i}"].value)

        charge_dict_list.append(charge_dict)
charge_df = pd.DataFrame.from_records(charge_dict_list)

# Read in the subtype data and add it to the output.
subtype_df = pd.read_csv(subtype_file)[["seqid", "subtype_ren"]]
subtype_df = subtype_df.rename(columns={"seqid": "seq.id", "subtype_ren": "subtype"})
output_df = pd.merge(output_df, subtype_df, how="left", on="seq.id")

# Read in the transmitted founder data, convert the founder string to the
# desired format, and add it to the output data.
founder_wb = load_workbook(filename=founder_file)
founder_ws = founder_wb.active
founder_dict_list = []
for i in range(2, founder_ws.max_row + 1):
    founder_dict = {}
    seq_id = founder_ws[f"A{i}"].value
    visitno = founder_ws[f"C{i}"].value
    if founder_ws[f"P{i}"].value.strip() == "Single":
        num_founders = "single"
    elif founder_ws[f"P{i}"].value.strip() == "Multi":
        num_founders = "multiple"
    elif founder_ws[f"P{i}"].value.strip() == "NA":
        num_founders = "NA"
    if ((seq_names_df["seq.id"] == seq_id) & (seq_names_df["visit"] == visitno)).any():
        founder_dict["seq.id"] = seq_id
        founder_dict["transmitted.founder.status"] = num_founders

        founder_dict_list.append(founder_dict)
founder_df = pd.DataFrame.from_records(founder_dict_list)
founder_df.drop_duplicates(keep="first", inplace=True)

output_df = pd.merge(output_df, founder_df, how="left", on="seq.id")

# Read in the HXB2 coordinate map.
with open(args.hxb2_file, "r") as hmf:
    file_lines = reader(hmf, delimiter="|")
    skip_header = next(file_lines)

    if skip_header is not None:
        for coordinates in file_lines:
            # The coordinate map dictionary exists in two forms:
            # 1) with the relative position as the key, in order to mark the
            #    unalignable positions in the insert sequences;
            # 2) with the absolute position as the key in order to create the
            #    correct variable name with the  sequence positions are matched.
            hxb2_map_dict[coordinates[1]] = int(coordinates[0]) - 1
            hxb2_abs_map_dict[int(coordinates[0]) - 1] = coordinates[1]

            # For the reference-independent sites, we need to have a mapping
            # that contains the HXB2 coordinates with something that will make
            # the insert positions sort properly if there are more than 26
            # insert positions.
            alpha_part = re.findall(r"\D+", coordinates[1])
            if alpha_part:
                alpha_string = "".join(alpha_part)
                number_part = re.findall(r"\d+", coordinates[1])
                number_string = "".join(number_part)
                if len(alpha_string) < 2:
                    underscore_separator = "_"
                else:
                    underscore_separator = ""
                hxb2_abs_sort_map_dict[int(coordinates[0]) - 1] = f"{number_string}{underscore_separator}{alpha_string}"
            else:
                hxb2_abs_sort_map_dict[int(coordinates[0]) - 1] = coordinates[1]

            # Get the absolute start and end positions for the regions of
            # interest that are contiguous. The positions are 1-based, so
            # subtract 1 from the start positions to get the 0-base.
            if coordinates[1] == f"{ENV_START_POS}":
                env_abs_start_pos = int(coordinates[0]) - 1
            elif coordinates[1] == f"{ENV_END_POS}":
                env_abs_end_pos = int(coordinates[0])
            if coordinates[1] == f"{GP120_START_POS}":
                gp120_abs_start_pos = int(coordinates[0]) - 1
            elif coordinates[1] == f"{GP120_END_POS}":
                gp120_abs_end_pos = int(coordinates[0])
            if coordinates[1] == f"{GP41_START_POS}":
                gp41_abs_start_pos = int(coordinates[0]) - 1
            elif coordinates[1] == f"{GP41_INSERT1_END_POS}":
                gp41_insert1_abs_end_pos = int(coordinates[0])
            elif coordinates[1] == f"{GP41_INSERT2_END_POS}":
                gp41_insert2_abs_end_pos = int(coordinates[0])
            elif coordinates[1] == f"{GP41_INSERT3_END_POS}":
                gp41_insert3_abs_end_pos = int(coordinates[0])
            if coordinates[1] == f"{INSERT1_START_POS}":
                insert1_abs_start_pos = int(coordinates[0]) - 1
            elif coordinates[1] == f"{INSERT1_END_POS}":
                insert1_abs_end_pos = int(coordinates[0])
            if coordinates[1] == f"{INSERT2_START_POS}":
                insert2_abs_start_pos = int(coordinates[0]) - 1
            elif coordinates[1] == f"{INSERT2_END_POS}":
                insert2_abs_end_pos = int(coordinates[0])
            if coordinates[1] == f"{INSERT3_START_POS}":
                insert3_abs_start_pos = int(coordinates[0]) - 1
            elif coordinates[1] == f"{INSERT3_END_POS}":
                insert3_abs_end_pos = int(coordinates[0])
            if coordinates[1] == f"{V2_START_POS}":
                v2_abs_start_pos = int(coordinates[0]) - 1
            elif coordinates[1] == f"{V2_END_POS}":
                v2_abs_end_pos = int(coordinates[0])
            if coordinates[1] == f"{V5_START_POS}":
                v5_abs_start_pos = int(coordinates[0]) - 1
            elif coordinates[1] == f"{V5_END_POS}":
                v5_abs_end_pos = int(coordinates[0])
            if coordinates[1] == f"{V1V2_START_POS}":
                v1v2_abs_start_pos = int(coordinates[0]) - 1
            elif coordinates[1] == f"{V1V2_END_POS}":
                v1v2_abs_end_pos = int(coordinates[0])

# Read in the V1V2 correlates sites, the V2 C1 ADCC epitope sites, the
# signature sites for C-clade bnAbs, and the cysteine conserved sites..
with open(v1v2_file, "r") as v1v2f:
    file_lines = reader(v1v2f)
    for coords in file_lines:
        v2_ab_list = v2_ab_list + coords

with open(adcc_file, "r") as adccf:
    file_lines = reader(adccf)
    for coords in file_lines:
        adcc_list = adcc_list + coords

with open(ant1428_pos_file, "r") as ant1428:
    file_lines = reader(ant1428)
    for coords in file_lines:
        antigen1428_list = antigen1428_list + coords

with open(cclade_file, "r") as ccf:
    file_lines = reader(ccf)
    for coords in file_lines:
        c_ab_list = c_ab_list + coords

with open(cysteine_conserved_file, "r") as cyscf:
    file_lines = reader(cyscf)
    for coords in file_lines:
        position = "".join(coords)
        conserved_cysteine_abs_pos_list.append(hxb2_map_dict[position])

sequon_pos_list = ["130", "156", "160", "234", "332", "362", "363", "392"]
for position in sequon_pos_list:
    sequon_abs_pos_list.append(hxb2_map_dict[position])

# Read in the HVTN505 CD4bs antibody contact sites and the sites of k-mers
# that overlap the CD4bs. The sites are already HXB2 aligned.
h505_cd4bs_df = pd.read_csv(hvtn505_site_file)
cd4bs_antibody_df = h505_cd4bs_df.loc[h505_cd4bs_df["set"] == "cd4bs_footprint"]
h505_cd4bs_antibody_list = cd4bs_antibody_df["hxb2.label"].tolist()
cd4bs_kmer_df = h505_cd4bs_df.loc[h505_cd4bs_df["set"] == "cd4bs_kmer_scan"]
h505_cd4bs_kmer_list = cd4bs_kmer_df["hxb2.label"].tolist()

# Create a list containing the Tier 1 designations for the appropriate
# positions.
tier1_list = {}
for i in range(env_abs_start_pos, env_abs_end_pos):
    if (((i >= v2_abs_start_pos) and (i < v2_abs_end_pos)) or (hxb2_abs_map_dict[i] in v2_ab_list)
        or (hxb2_abs_map_dict[i] in adcc_list) or (hxb2_abs_map_dict[i] in c_ab_list)):
        tier1_list[i] = ".tier1"
    else:
        tier1_list[i] = ""

# Create a list containing the Tier 1 designations for the appropriate
# a1428 positions.
tier1_a1428_list = {}
for a1428_site in antigen1428_list:
    if (((a1428_map_dict[a1428_site] >= v2_abs_start_pos) and (a1428_map_dict[a1428_site] <= v2_abs_end_pos))
        or (a1428_site in v2_ab_list) or (a1428_site in adcc_list) or (a1428_site in c_ab_list)):
        tier1_a1428_list[a1428_map_dict[a1428_site]] = ".tier1"
    else:
        tier1_a1428_list[a1428_map_dict[a1428_site]] = ""

# Read in the unalignable sites, and use the reference map to change the
# characters that are in this position in the insert strings to a "!".
insert1_match = insert1_seq.copy()
insert2_match = insert2_seq.copy()
insert3_match = insert3_seq.copy()
insert4_match = insert4_seq.copy()
with open(unalignable_file, "r") as uf:
    file_lines = reader(uf)
    for coords in file_lines:
        insert1_match[int(hxb2_map_dict[coords[0]])] = "!"
        insert2_match[int(hxb2_map_dict[coords[0]])] = "!"
        insert3_match[int(hxb2_map_dict[coords[0]])] = "!"
        if coords[0] in antigen1428_list:
            insert4_match[int(a1428_map_dict[coords[0]])] = "!"

# Create the Hamming distance table.
hamming_df = pd.read_csv(hamming_file)
hamming_df.set_index("aa1", inplace=True)

# Match the sequences with the inserts by position, calculate Hamming
# distances, calculate lengths, and determine sequons.
for seq_id in seq_dict:
    cysteine_dict = {}
    cysteine_dict["seq.id"] = seq_id
    cysteine_dict["cysteine.count"] = 0
    match1_dict = {}
    match1_dict["seq.id"] = seq_id
    match2_dict = {}
    hamming2_dict = {}
    match2_dict["seq.id"] = seq_id
    match3_dict = {}
    match3_dict["seq.id"] = seq_id
    match4_dict = {}
    match4_dict["seq.id"] = seq_id
    non_conserved_cysteine_dict = {}
    non_conserved_cysteine_dict["seq.id"] = seq_id
    non_conserved_cysteine_dict["nonconserved.cysteine.count"] = 0
    reference_independent_dict = {}
    reference_independent_dict["seq.id"] = seq_id
    hamming_dict = {}
    hamming_dict["seq.id"] = seq_id
    hamming_dict["hdist.zspace.mos1.v2"] = 0
    hamming_dict["hdist.zspace.mos2.v2"] = 0
    hamming_dict["hdist.zspace.c97za.v2"] = 0
    hamming_dict["hdist.zspace.mos1.v2_ab"] = 0
    hamming_dict["hdist.zspace.mos2.v2_ab"] = 0
    hamming_dict["hdist.zspace.c97za.v2_ab"] = 0
    hamming_dict["hdist.zspace.mos1.adcc"] = 0
    hamming_dict["hdist.zspace.mos2.adcc"] = 0
    hamming_dict["hdist.zspace.c97za.adcc"] = 0
    hamming_dict["hdist.zspace.mos1.c_ab"] = 0
    hamming_dict["hdist.zspace.mos2.c_ab"] = 0
    hamming_dict["hdist.zspace.c97za.c_ab"] = 0
    hamming_dict["hdist.zspace.mos1.c_ab_no364"] = 0
    hamming_dict["hdist.zspace.mos2.c_ab_no364"] = 0
    hamming_dict["hdist.zspace.c97za.c_ab_no364"] = 0
    hamming_dict["hdist.zspace.mos1.gp41"] = 0
    hamming_dict["hdist.zspace.mos2.gp41"] = 0
    hamming_dict["hdist.zspace.c97za.gp41"] = 0
    hamming_dict["hdist.zspace.mos1.gp120"] = 0
    hamming_dict["hdist.zspace.mos2.gp120"] = 0
    hamming_dict["hdist.zspace.c97za.gp120"] = 0
    hamming_dict["hdist.zspace.mos1.v5"] = 0
    hamming_dict["hdist.zspace.mos2.v5"] = 0
    hamming_dict["hdist.zspace.c97za.v5"] = 0
    hamming_dict["hdist.zspace.1428v1v2"] = 0
    hamming_dict["hdist.zspace.mos1.hvtn505.cd4bs.antibody"] = 0
    hamming_dict["hdist.zspace.mos1.hvtn505.cd4bs.kmer"] = 0
    hamming_dict["hdist.zspace.mos2.hvtn505.cd4bs.antibody"] = 0
    hamming_dict["hdist.zspace.mos2.hvtn505.cd4bs.kmer"] = 0
    hamming_dict["hdist.zspace.c97za.hvtn505.cd4bs.antibody"] = 0
    hamming_dict["hdist.zspace.c97za.hvtn505.cd4bs.kmer"] = 0
    length_dict = {}
    length_dict["seq.id"] = seq_id
    length_dict["length.v1v2"] = 0
    length_dict["length.v5"] = 0
    sequon_dict = {}
    sequon_dict["seq.id"] = seq_id
    seq_sequon_list = sequon_abs_pos_list.copy()
    curr_sequon_pos = seq_sequon_list.pop(0)
    region_sequon_dict = {}
    region_sequon_dict["seq.id"] = seq_id

    # Cycle through the gp120 region and match the amino acids in the sequences
    # to those in the same positions in the insert sequences.
    for i in range(env_abs_start_pos, env_abs_end_pos):
        # Since the unalignable positions are the same for all of the inserts,
        # it is only necessary to check one of them. Recall that we had earlier
        # in the program set the value of the unalignable positions to the
        # character "!". Since insert2 (Ad26.Mos2S.Env) is the longest insert
        # use it to check for unalignable positions.
        if insert2_match[i] != "!":
            if insert1_abs_start_pos <= i < insert1_abs_end_pos:
                if seq_dict[seq_id][i] == insert1_match[i]:
                    match1_dict[f"hxb2.{hxb2_abs_map_dict[i]}.1mer.mos1.match{tier1_list[i]}"] = "1"
                else:
                    match1_dict[f"hxb2.{hxb2_abs_map_dict[i]}.1mer.mos1.match{tier1_list[i]}"] = "0"
                # Calculate the Hamming distances.
                if gp120_abs_start_pos <= i < gp120_abs_end_pos:
                    hamming_dict["hdist.zspace.mos1.gp120"] = hamming_dict["hdist.zspace.mos1.gp120"] + hamming_df.loc[seq_dict[seq_id][i], insert1_seq[i]]
                elif gp41_abs_start_pos <= i < gp41_insert1_abs_end_pos:
                    hamming_dict["hdist.zspace.mos1.gp41"] = hamming_dict["hdist.zspace.mos1.gp41"] + hamming_df.loc[seq_dict[seq_id][i], insert1_seq[i]]

            if insert2_abs_start_pos <= i < insert2_abs_end_pos:
                if seq_dict[seq_id][i] == insert2_match[i]:
                    match2_dict[f"hxb2.{hxb2_abs_map_dict[i]}.1mer.mos2.match{tier1_list[i]}"] = "1"
                else:
                    match2_dict[f"hxb2.{hxb2_abs_map_dict[i]}.1mer.mos2.match{tier1_list[i]}"] = "0"
                # Calculate the Hamming distances.
                if gp120_abs_start_pos <= i < gp120_abs_end_pos:
                    hamming_dict["hdist.zspace.mos2.gp120"] = hamming_dict["hdist.zspace.mos2.gp120"] + hamming_df.loc[seq_dict[seq_id][i], insert2_seq[i]]
                elif gp41_abs_start_pos <= i < gp41_insert2_abs_end_pos:
                    hamming_dict["hdist.zspace.mos2.gp41"] = hamming_dict["hdist.zspace.mos2.gp41"] + hamming_df.loc[seq_dict[seq_id][i], insert2_seq[i]]

            if insert3_abs_start_pos <= i < insert3_abs_end_pos:
                if seq_dict[seq_id][i] == insert3_match[i]:
                    match3_dict[f"hxb2.{hxb2_abs_map_dict[i]}.1mer.c97za.match{tier1_list[i]}"] = "1"
                else:
                    match3_dict[f"hxb2.{hxb2_abs_map_dict[i]}.1mer.c97za.match{tier1_list[i]}"] = "0"
                # Calculate the Hamming distances.
                if gp120_abs_start_pos <= i < gp120_abs_end_pos:
                    hamming_dict["hdist.zspace.c97za.gp120"] = hamming_dict["hdist.zspace.c97za.gp120"] + hamming_df.loc[seq_dict[seq_id][i], insert3_seq[i]]
                elif gp41_abs_start_pos <= i < gp41_insert3_abs_end_pos:
                    hamming_dict["hdist.zspace.c97za.gp41"] = hamming_dict["hdist.zspace.c97za.gp41"] + hamming_df.loc[seq_dict[seq_id][i], insert3_seq[i]]

            if (i in a1428_abs_map_dict) and (a1428_abs_map_dict[i] in antigen1428_list):
                if seq_dict[seq_id][hxb2_map_dict[a1428_abs_map_dict[i]]] == insert4_match[i]:
                    match4_dict[f"hxb2.{a1428_abs_map_dict[i]}.1mer.1428v1v2.match{tier1_a1428_list[i]}"] = "1"
                else:
                    match4_dict[f"hxb2.{a1428_abs_map_dict[i]}.1mer.1428v1v2.match{tier1_a1428_list[i]}"] = "0"

            # Determine the reference independent marks.
            if seq_dict[seq_id][i] == "-":
                seq_aa = "gap"
            else:
                seq_aa = seq_dict[seq_id][i]
            if i < 9:
                dict_key = f"hxb2.00{hxb2_abs_sort_map_dict[i]}.is.{seq_aa}{tier1_list[i]}"
            # According to the HXB2 map, triple digits start at position 119.
            # This would be 118 since the iterator is 0-based.
            elif 9 <= i < 118:
                dict_key = f"hxb2.0{hxb2_abs_sort_map_dict[i]}.is.{seq_aa}{tier1_list[i]}"
            else:
                dict_key = f"hxb2.{hxb2_abs_sort_map_dict[i]}.is.{seq_aa}{tier1_list[i]}"
            reference_independent_dict[dict_key] = 1

            # Check if there is a cysteine (C) at the current position. Only do
            # this for gp160, which starts at the beginning of gp120 and goes
            # to the end of the longest insert (Ad26.Mos2S.Env).
            if i >= gp120_abs_start_pos:
                if seq_dict[seq_id][i] == "C":
                    cysteine_dict["cysteine.count"] += 1

            # Check if there is a cysteine (C) at the current position if the
            # position is not on the conserved positions list. Only do this
            # for gp160, which starts at the beginning of gp120 and goes to
            # the end of the longest insert (Ad26.Mos2S.Env).
            if (i >= gp120_abs_start_pos) and (i not in conserved_cysteine_abs_pos_list):
                if seq_dict[seq_id][i] == "C":
                    non_conserved_cysteine_dict["nonconserved.cysteine.count"] += 1

            # For the regions that don't have contiguous positions, the
            # positions previously read from the files are relative, so compare
            # the relative position referenced by the current absolute position
            # counter.
            if hxb2_abs_map_dict[i] in v2_ab_list:
                hamming_dict["hdist.zspace.mos1.v2_ab"] = hamming_dict["hdist.zspace.mos1.v2_ab"] + hamming_df.loc[seq_dict[seq_id][i], insert1_seq[i]]
                hamming_dict["hdist.zspace.mos2.v2_ab"] = hamming_dict["hdist.zspace.mos2.v2_ab"] + hamming_df.loc[seq_dict[seq_id][i], insert2_seq[i]]
                hamming_dict["hdist.zspace.c97za.v2_ab"] = hamming_dict["hdist.zspace.c97za.v2_ab"] + hamming_df.loc[seq_dict[seq_id][i], insert3_seq[i]]

            if hxb2_abs_map_dict[i] in adcc_list:
                hamming_dict["hdist.zspace.mos1.adcc"] = hamming_dict["hdist.zspace.mos1.adcc"] + hamming_df.loc[seq_dict[seq_id][i], insert1_seq[i]]
                hamming_dict["hdist.zspace.mos2.adcc"] = hamming_dict["hdist.zspace.mos2.adcc"] + hamming_df.loc[seq_dict[seq_id][i], insert2_seq[i]]
                hamming_dict["hdist.zspace.c97za.adcc"] = hamming_dict["hdist.zspace.c97za.adcc"] + hamming_df.loc[seq_dict[seq_id][i], insert3_seq[i]]

            if hxb2_abs_map_dict[i] in c_ab_list:
                hamming_dict["hdist.zspace.mos1.c_ab"] = hamming_dict["hdist.zspace.mos1.c_ab"] + hamming_df.loc[seq_dict[seq_id][i], insert1_seq[i]]
                hamming_dict["hdist.zspace.mos2.c_ab"] = hamming_dict["hdist.zspace.mos2.c_ab"] + hamming_df.loc[seq_dict[seq_id][i], insert2_seq[i]]
                hamming_dict["hdist.zspace.c97za.c_ab"] = hamming_dict["hdist.zspace.c97za.c_ab"] + hamming_df.loc[seq_dict[seq_id][i], insert3_seq[i]]
                if hxb2_abs_map_dict[i] != "364":
                    hamming_dict["hdist.zspace.mos1.c_ab_no364"] = hamming_dict["hdist.zspace.mos1.c_ab_no364"] + hamming_df.loc[seq_dict[seq_id][i], insert1_seq[i]]
                    hamming_dict["hdist.zspace.mos2.c_ab_no364"] = hamming_dict["hdist.zspace.mos2.c_ab_no364"] + hamming_df.loc[seq_dict[seq_id][i], insert2_seq[i]]
                    hamming_dict["hdist.zspace.c97za.c_ab_no364"] = hamming_dict["hdist.zspace.c97za.c_ab_no364"] + hamming_df.loc[seq_dict[seq_id][i], insert3_seq[i]]

            if hxb2_abs_map_dict[i] in antigen1428_list:
                a1428_map_pos = a1428_map_dict[hxb2_abs_map_dict[i]]
                hamming_dict["hdist.zspace.1428v1v2"] = hamming_dict["hdist.zspace.1428v1v2"] + hamming_df.loc[seq_dict[seq_id][i], insert4_seq[a1428_map_pos]]

            if (i >= v2_abs_start_pos) and (i < v2_abs_end_pos):
                hamming_dict["hdist.zspace.mos1.v2"] = hamming_dict["hdist.zspace.mos1.v2"] + hamming_df.loc[seq_dict[seq_id][i], insert1_seq[i]]
                hamming_dict["hdist.zspace.mos2.v2"] = hamming_dict["hdist.zspace.mos2.v2"] + hamming_df.loc[seq_dict[seq_id][i], insert2_seq[i]]
                hamming_dict["hdist.zspace.c97za.v2"] = hamming_dict["hdist.zspace.c97za.v2"] + hamming_df.loc[seq_dict[seq_id][i], insert3_seq[i]]

            if (i >= v5_abs_start_pos) and (i < v5_abs_end_pos):
                hamming_dict["hdist.zspace.mos1.v5"] = hamming_dict["hdist.zspace.mos1.v5"] + hamming_df.loc[seq_dict[seq_id][i], insert1_seq[i]]
                hamming_dict["hdist.zspace.mos2.v5"] = hamming_dict["hdist.zspace.mos2.v5"] + hamming_df.loc[seq_dict[seq_id][i], insert2_seq[i]]
                hamming_dict["hdist.zspace.c97za.v5"] = hamming_dict["hdist.zspace.c97za.v5"] + hamming_df.loc[seq_dict[seq_id][i], insert3_seq[i]]
            if hxb2_abs_map_dict[i] in h505_cd4bs_antibody_list:
                hamming_dict["hdist.zspace.mos1.hvtn505.cd4bs.antibody"] = hamming_dict["hdist.zspace.mos1.hvtn505.cd4bs.antibody"] + hamming_df.loc[seq_dict[seq_id][i], insert1_seq[i]]
                hamming_dict["hdist.zspace.mos2.hvtn505.cd4bs.antibody"] = hamming_dict["hdist.zspace.mos2.hvtn505.cd4bs.antibody"] + hamming_df.loc[seq_dict[seq_id][i], insert2_seq[i]]
                hamming_dict["hdist.zspace.c97za.hvtn505.cd4bs.antibody"] = hamming_dict["hdist.zspace.c97za.hvtn505.cd4bs.antibody"] + hamming_df.loc[seq_dict[seq_id][i], insert3_seq[i]]
            if hxb2_abs_map_dict[i] in h505_cd4bs_kmer_list:
                hamming_dict["hdist.zspace.mos1.hvtn505.cd4bs.kmer"] = hamming_dict["hdist.zspace.mos1.hvtn505.cd4bs.kmer"] + hamming_df.loc[seq_dict[seq_id][i], insert1_seq[i]]
                hamming_dict["hdist.zspace.mos2.hvtn505.cd4bs.kmer"] = hamming_dict["hdist.zspace.mos2.hvtn505.cd4bs.kmer"] + hamming_df.loc[seq_dict[seq_id][i], insert2_seq[i]]
                hamming_dict["hdist.zspace.c97za.hvtn505.cd4bs.kmer"] = hamming_dict["hdist.zspace.c97za.hvtn505.cd4bs.kmer"] + hamming_df.loc[seq_dict[seq_id][i], insert3_seq[i]]


        # Calculate the lengths (i.e. number of residues) of the V1V2 and V5
        # regions.
        if (i >= v1v2_abs_start_pos) and (i < v1v2_abs_end_pos) and (seq_dict[seq_id][i] != "-"):
            length_dict["length.v1v2"] += 1
        if (i >= v5_abs_start_pos) and (i < v5_abs_end_pos) and (seq_dict[seq_id][i] != "-"):
            length_dict["length.v5"] += 1

        # Determine if there are sequons that start at the specified positions.
        # The sequon motif is [N][not P][S or T]. Gaps and the marker for
        # unalignable sites are excluded. The regex expression below translates to:
        # 1) starts with N at the designated position
        # 2) followed by a character that is not P
        # 3) followed by an S or a T
        if (i == curr_sequon_pos):
            curr_sequon = seq_dict[seq_id][i]
            add_index = 1
            while len(curr_sequon) < 3:
                if seq_dict[seq_id][i + add_index] != "-":
                    curr_sequon += seq_dict[seq_id][i + add_index]
                add_index += 1
            is_match = re.match(r"N[^P][ST]", curr_sequon)
            if is_match:
                sequon_dict[f"hxb2.{hxb2_abs_map_dict[curr_sequon_pos]}.is.sequon.tier1"] = "1"
            else:
                sequon_dict[f"hxb2.{hxb2_abs_map_dict[curr_sequon_pos]}.is.sequon.tier1"] = "0"
            if seq_sequon_list:
                curr_sequon_pos = seq_sequon_list.pop(0)

    # Determine how many sequons there are in the specified regions.
    v1v2_region = "".join(seq_dict[seq_id][v1v2_abs_start_pos : v1v2_abs_end_pos])
    v1v2_region = v1v2_region.replace("-", "")
    v1v2_region = v1v2_region.replace("!", "")
    v1v2_sequons = re.findall(r"(?=(N[^P][ST]))", v1v2_region)
    if v1v2_sequons:
        region_sequon_dict["num.sequons.v1v2"] = str(len(v1v2_sequons))
    else:
        region_sequon_dict["num.sequons.v1v2"] = "0"

    v5_region = "".join(seq_dict[seq_id][v5_abs_start_pos : v5_abs_end_pos])
    v5_region = v5_region.replace("-", "")
    v5_region = v5_region.replace("!", "")
    v5_sequons = re.findall(r"(?=(N[^P][ST]))", v5_region)
    if v5_sequons:
        region_sequon_dict["num.sequons.v5"] = str(len(v5_sequons))
    else:
        region_sequon_dict["num.sequons.v5"] = "0"
        
    seq_match_mos1_list.append(match1_dict)
    seq_match_mos2_list.append(match2_dict)
    seq_match_c97za_list.append(match3_dict)
    seq_match_1428_list.append(match4_dict)

    reference_independent_list.append(reference_independent_dict)

    # Round the Hamming distances to 4 decimal places.
    hamming_dict["hdist.zspace.mos1.gp120"] = round(hamming_dict["hdist.zspace.mos1.gp120"], 4)
    hamming_dict["hdist.zspace.mos2.gp120"] = round(hamming_dict["hdist.zspace.mos2.gp120"], 4)
    hamming_dict["hdist.zspace.c97za.gp120"] = round(hamming_dict["hdist.zspace.c97za.gp120"], 4)
    hamming_dict["hdist.zspace.mos1.v2_ab"] = round(hamming_dict["hdist.zspace.mos1.v2_ab"], 4)
    hamming_dict["hdist.zspace.mos2.v2_ab"] = round(hamming_dict["hdist.zspace.mos2.v2_ab"], 4)
    hamming_dict["hdist.zspace.c97za.v2_ab"] = round(hamming_dict["hdist.zspace.c97za.v2_ab"], 4)
    hamming_dict["hdist.zspace.mos1.adcc"] = round(hamming_dict["hdist.zspace.mos1.adcc"], 4)
    hamming_dict["hdist.zspace.mos2.adcc"] = round(hamming_dict["hdist.zspace.mos2.adcc"], 4)
    hamming_dict["hdist.zspace.c97za.adcc"] = round(hamming_dict["hdist.zspace.c97za.adcc"] , 4)
    hamming_dict["hdist.zspace.mos1.c_ab"] = round(hamming_dict["hdist.zspace.mos1.c_ab"], 4)
    hamming_dict["hdist.zspace.mos2.c_ab"] = round(hamming_dict["hdist.zspace.mos2.c_ab"], 4)
    hamming_dict["hdist.zspace.c97za.c_ab"] = round(hamming_dict["hdist.zspace.c97za.c_ab"] , 4)
    hamming_dict["hdist.zspace.mos1.c_ab_no364"] = round(hamming_dict["hdist.zspace.mos1.c_ab_no364"], 4)
    hamming_dict["hdist.zspace.mos2.c_ab_no364"] = round(hamming_dict["hdist.zspace.mos2.c_ab_no364"], 4)
    hamming_dict["hdist.zspace.c97za.c_ab_no364"] = round(hamming_dict["hdist.zspace.c97za.c_ab_no364"] , 4)
    hamming_dict["hdist.zspace.1428v1v2"] = round(hamming_dict["hdist.zspace.1428v1v2"], 4)
    hamming_dict["hdist.zspace.mos1.v2"] = round(hamming_dict["hdist.zspace.mos1.v2"], 4)
    hamming_dict["hdist.zspace.mos2.v2"] = round(hamming_dict["hdist.zspace.mos2.v2"], 4)
    hamming_dict["hdist.zspace.c97za.v2"] = round(hamming_dict["hdist.zspace.c97za.v2"] , 4)
    hamming_dict["hdist.zspace.mos1.v5"] = round(hamming_dict["hdist.zspace.mos1.v5"], 4)
    hamming_dict["hdist.zspace.mos2.v5"] = round(hamming_dict["hdist.zspace.mos2.v5"], 4)
    hamming_dict["hdist.zspace.c97za.v5"] = round(hamming_dict["hdist.zspace.c97za.v5"] , 4)
    hamming_dict["hdist.zspace.mos1.hvtn505.cd4bs.antibody"] = round(hamming_dict["hdist.zspace.mos1.hvtn505.cd4bs.antibody"], 4)
    hamming_dict["hdist.zspace.mos1.hvtn505.cd4bs.kmer"] = round(hamming_dict["hdist.zspace.mos1.hvtn505.cd4bs.kmer"], 4) 
    hamming_dict["hdist.zspace.mos2.hvtn505.cd4bs.antibody"] = round(hamming_dict["hdist.zspace.mos2.hvtn505.cd4bs.antibody"], 4)
    hamming_dict["hdist.zspace.mos2.hvtn505.cd4bs.kmer"] = round(hamming_dict["hdist.zspace.mos2.hvtn505.cd4bs.kmer"], 4) 
    hamming_dict["hdist.zspace.c97za.hvtn505.cd4bs.antibody"] = round(hamming_dict["hdist.zspace.c97za.hvtn505.cd4bs.antibody"], 4)
    hamming_dict["hdist.zspace.c97za.hvtn505.cd4bs.kmer"] = round(hamming_dict["hdist.zspace.c97za.hvtn505.cd4bs.kmer"], 4) 
    hamming_list.append(hamming_dict)

    length_dict["length.v1v2"] = str(length_dict["length.v1v2"])
    length_dict["length.v5"] = str(length_dict["length.v5"])
    length_list.append(length_dict)
    sequon_list.append(sequon_dict)
    region_sequon_list.append(region_sequon_dict)
    cysteine_count_list.append(cysteine_dict)
    non_conserved_cysteine_count_list.append(non_conserved_cysteine_dict)

# Create dataframes from the lists and then merge the dataframes to create
# the final output dataframe.
match1_df = pd.DataFrame.from_records(seq_match_mos1_list)
output_df = pd.merge(output_df, match1_df, how="left", on="seq.id")
match2_df = pd.DataFrame.from_records(seq_match_mos2_list)
output_df = pd.merge(output_df, match2_df, how="left", on="seq.id")
match3_df = pd.DataFrame.from_records(seq_match_c97za_list)
output_df = pd.merge(output_df, match3_df, how="left", on="seq.id")
match4_df = pd.DataFrame.from_records(seq_match_1428_list)
output_df = pd.merge(output_df, match4_df, how="left", on="seq.id")

reference_independent_df = pd.DataFrame.from_records(reference_independent_list)
reference_independent_df = reference_independent_df.reindex(sorted(reference_independent_df.columns), axis=1)
reference_independent_df.fillna(0, inplace=True)
reference_independent_df.columns = reference_independent_df.columns.str.replace(r"\.00", r".", regex=True)
reference_independent_df.columns = reference_independent_df.columns.str.replace(r"\.0", r".", regex=True)
reference_independent_df.columns = reference_independent_df.columns.str.replace(r"_", r"", regex=True)
output_df = pd.merge(output_df, reference_independent_df, how="left", on="seq.id")

hamdist_df = pd.DataFrame.from_records(hamming_list)
output_df = pd.merge(output_df, hamdist_df, how="left", on="seq.id")
length_df = pd.DataFrame.from_records(length_list)
output_df = pd.merge(output_df, length_df, how="left", on="seq.id")
output_df = pd.merge(output_df, charge_df, how="left", on="seq.id")
sequon_df = pd.DataFrame.from_records(sequon_list)
output_df = pd.merge(output_df, sequon_df, how="left", on="seq.id")
region_sequon_df = pd.DataFrame.from_records(region_sequon_list)
output_df = pd.merge(output_df, region_sequon_df, how="left", on="seq.id")
cysteine_df = pd.DataFrame.from_records(cysteine_count_list)
output_df = pd.merge(output_df, cysteine_df, how="left", on="seq.id")
non_conserved_cysteine_df = pd.DataFrame.from_records(non_conserved_cysteine_count_list)
output_df = pd.merge(output_df, non_conserved_cysteine_df, how="left", on="seq.id")
output_df.fillna("NA", inplace=True)

output_df.to_csv(args.out_file, index=False)
